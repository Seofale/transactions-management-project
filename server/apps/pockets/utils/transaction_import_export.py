from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.request import Request

from ...pockets import models
from ...pockets.models import querysets
from ..serializers import TransactionCreateSerializer
from config import settings


def save_transactions_as_xlsx(transactions: querysets.TransactionQuerySet) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Транзакции'

    worksheet.cell(row=1, column=1, value='Дата')
    worksheet.cell(row=1, column=2, value='Сумма')
    worksheet.cell(row=1, column=3, value='Категория')
    worksheet.cell(row=1, column=4, value='Тип операции')

    row = 2
    column = 1

    for transaction in transactions:

        if transaction.category:
            category = transaction.category.name
        else:
            category = None

        worksheet.cell(row=row, column=column, value=transaction.transaction_date)
        worksheet.cell(row=row, column=column+1, value=transaction.amount)
        worksheet.cell(row=row, column=column+2, value=category)
        worksheet.cell(row=row, column=column+3, value=transaction.transaction_type)

        row += 1

    workbook.save(f'./{settings.MEDIA_URL}/transaction_exports/{transactions.first().user}--{timezone.now()}.xlsx')

    return workbook


def create_transactions_by_xlsx_file(file_obj, request: Request):
    errors = []

    try:
        workbook = load_workbook(file_obj)
    except Exception as e:
        return str(e)

    validated_transactions = []

    work_sheet = workbook.active

    row_count = work_sheet.max_row

    for row in work_sheet.iter_rows(min_row=2, min_col=1, max_row=row_count, max_col=4):
        transaction_date, amount, category_name, transaction_type = row

        if category_name.value:
            category, _ = models.TransactionCategory.objects.get_or_create(
                user=request.user,
                name=category_name.value
            )
        else:
            category = None

        data = {
            'transaction_date': transaction_date.value.date(),
            'amount': amount.value,
            'category': category.id if category else None,
            'transaction_type': transaction_type.value
        }

        serializer = TransactionCreateSerializer(data=data, context={'request': request})
        serializer.is_valid()
        serializer.validated_data.update({'user': request.user})

        if serializer.errors:
            errors.append(serializer.errors)

        transaction = models.Transaction(**serializer.validated_data)
        validated_transactions.append(transaction)

    if not errors:
        models.Transaction.objects.bulk_create(validated_transactions)

    return errors
